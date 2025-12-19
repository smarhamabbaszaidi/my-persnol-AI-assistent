# phoenix_ai_with_sidebar.py
import os
import sys
import subprocess
import speech_recognition as sr
import pyttsx3
import webbrowser
import requests
import datetime
import openpyxl
from openpyxl import Workbook
import threading
import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox, filedialog
import psutil
import wikipedia
import time
from PIL import Image, ImageTk
import sv_ttk  # For modern theme support
##############################
####you want api key here#####
# Initialize APIs (kept exactly as you provided)
GEMINI_API_KEY = "Paste your api here"
GEMINI_API_URL = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent?key={GEMINI_API_KEY}"
DEEPSEEK_API_URL = "https://api.deepseek.com/v1/chat/completions"
DEEPSEEK_API_KEY = "sk-d63c79e8e06140009a7d44f8a00984e5"  # Replace with actual key if needed

# Initialize voice engine
engine = pyttsx3.init()
voices = engine.getProperty('voices')
if len(voices) > 1:
    engine.setProperty('voice', voices[1].id)  # Female voice if available
engine.setProperty('rate', 180)

# Password
PASSWORD = "XENO2125"

# App paths
APP_PATHS = {
    "paint": "mspaint",
    "chrome": "chrome",
    "youtube": "https://youtube.com",
    "google": "https://google.com",
    "microsoft store": "ms-windows-store:",
    "notepad": "notepad",
    "calculator": "calc",
    "word": "winword",
    "excel": "excel",
    "powerpoint": "powerpnt",
    "file manager": "explorer",
    "photos": "ms-photos:",
    "settings": "ms-settings:",
    "spotify": "spotify",
    "discord": "discord",
    "steam": "steam",
    "command prompt": "cmd",
    "powershell": "powershell"
}


class PhoenixAI:
    def __init__(self, root):
        self.root = root
        self.root.title("Phoenix AI Assistant")
        self.root.geometry("1200x800")
        self.root.minsize(1000, 700)

        # Theme
        sv_ttk.set_theme("dark")

        # Style tweaks
        self.style = ttk.Style()
        self.style.configure('Title.TLabel', font=('Helvetica', 24, 'bold'))
        self.style.configure('Status.TLabel', font=('Helvetica', 12))
        self.style.configure('Accent.TButton', font=('Helvetica', 11, 'bold'))

        # Chat history store (list of tuples: (sender, text, iso timestamp))
        self.chat_history = []

        self.setup_ui()

        # Speech recognizer
        self.recognizer = sr.Recognizer()
        self.microphone = sr.Microphone()
        with self.microphone as source:
            self.recognizer.adjust_for_ambient_noise(source)

        self.listening = False
        self.authenticated = False
        self.current_excel_file = None
        self.excel_mode = False

        # Greet and ask for password
        self.speak("Hello Master. Please enter the password to activate Phoenix.")

    def setup_ui(self):
        main_container = ttk.Frame(self.root, padding="10")
        main_container.pack(fill=tk.BOTH, expand=True)

        # Header
        header_frame = ttk.Frame(main_container)
        header_frame.pack(fill=tk.X, pady=(0, 15))

        title_frame = ttk.Frame(header_frame)
        title_frame.pack(side=tk.LEFT)

        self.logo_label = ttk.Label(title_frame, text="🔥", font=("Helvetica", 28))
        self.logo_label.pack(side=tk.LEFT, padx=(0, 10))

        self.title_label = ttk.Label(title_frame, text="PHOENIX AI", style='Title.TLabel')
        self.title_label.pack(side=tk.LEFT)

        status_frame = ttk.Frame(header_frame)
        status_frame.pack(side=tk.RIGHT)

        self.status_var = tk.StringVar(value="Inactive")
        status_label = ttk.Label(status_frame, textvariable=self.status_var, style='Status.TLabel')
        status_label.pack(side=tk.RIGHT)
        ttk.Label(status_frame, text="Status: ", style='Status.TLabel').pack(side=tk.RIGHT)

        stats_frame = ttk.Frame(header_frame)
        stats_frame.pack(side=tk.RIGHT, padx=20)
        self.cpu_var = tk.StringVar(value="CPU: --%")
        self.mem_var = tk.StringVar(value="Memory: --%")
        ttk.Label(stats_frame, textvariable=self.cpu_var, style='Status.TLabel').pack(anchor=tk.E)
        ttk.Label(stats_frame, textvariable=self.mem_var, style='Status.TLabel').pack(anchor=tk.E)

        # Content area uses PanedWindow so we can have a resizable previous-chats panel
        content_pane = ttk.Panedwindow(main_container, orient=tk.HORIZONTAL)
        content_pane.pack(fill=tk.BOTH, expand=True)

        # Left control panel
        left_panel = ttk.Frame(content_pane, width=250)
        left_panel.pack_propagate(False)
        content_pane.add(left_panel, weight=1)

        # Authentication
        auth_frame = ttk.LabelFrame(left_panel, text="Authentication", padding=10)
        auth_frame.pack(fill=tk.X, pady=(0, 10))
        ttk.Label(auth_frame, text="Password:").pack(anchor=tk.W)
        self.pass_entry = ttk.Entry(auth_frame, show="*", width=20)
        self.pass_entry.pack(fill=tk.X, pady=(5, 10))
        self.pass_entry.bind("<Return>", lambda e: self.authenticate())
        self.auth_btn = ttk.Button(auth_frame, text="Authenticate", command=self.authenticate, style='Accent.TButton')
        self.auth_btn.pack(fill=tk.X)

        # Controls
        controls_frame = ttk.LabelFrame(left_panel, text="Controls", padding=10)
        controls_frame.pack(fill=tk.X, pady=(0, 10))
        self.listen_btn = ttk.Button(controls_frame, text="Start Listening", command=self.toggle_listening)
        self.listen_btn.pack(fill=tk.X, pady=2)
        self.excel_btn = ttk.Button(controls_frame, text="Excel Mode: Off", command=self.toggle_excel_mode)
        self.excel_btn.pack(fill=tk.X, pady=2)
        self.file_btn = ttk.Button(controls_frame, text="Open File", command=self.open_file_dialog)
        self.file_btn.pack(fill=tk.X, pady=2)

        # Quick actions
        actions_frame = ttk.LabelFrame(left_panel, text="Quick Actions", padding=10)
        actions_frame.pack(fill=tk.X)
        actions = ["Time", "Date", "Weather", "News", "Email", "Calendar"]
        for i, action in enumerate(actions):
            btn = ttk.Button(actions_frame, text=action,
                             command=lambda a=action: self.quick_action(a.lower()))
            btn.grid(row=i // 2, column=i % 2, padx=2, pady=2, sticky="ew")
        actions_frame.columnconfigure(0, weight=1)
        actions_frame.columnconfigure(1, weight=1)

        # Main (center) area: we'll create communication log + input box
        center_panel = ttk.Frame(content_pane)
        content_pane.add(center_panel, weight=4)

        # Communication Log frame
        console_frame = ttk.LabelFrame(center_panel, text="Communication Log", padding=10)
        console_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))

        self.console = scrolledtext.ScrolledText(console_frame, width=60, height=15,
                                                font=("Consolas", 10), bg="#1a1a2e", fg="#00ff00",
                                                insertbackground="#00ff00", relief="flat")
        self.console.pack(fill=tk.BOTH, expand=True)
        self.console.config(state=tk.DISABLED)

        # Visualization area underneath console
        viz_frame = ttk.Frame(center_panel, height=80)
        viz_frame.pack(fill=tk.X)
        viz_frame.pack_propagate(False)
        self.canvas = tk.Canvas(viz_frame, bg="#0f0f23", highlightthickness=0)
        self.canvas.pack(fill=tk.BOTH, expand=True)

        # Input area (text box + send + mic button) - mic at the corner
        input_frame = ttk.Frame(center_panel, padding=(0, 6))
        input_frame.pack(fill=tk.X)

        # Large Text widget for multi-line input
        self.input_text = tk.Text(input_frame, height=3, font=("Segoe UI", 11))
        self.input_text.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 6))
        self.input_text.bind("<Control-Return>", lambda e: self.on_send_text())  # Ctrl+Enter to send

        # Right-side small frame with buttons
        input_buttons = ttk.Frame(input_frame)
        input_buttons.pack(side=tk.RIGHT, fill=tk.Y)

        self.send_btn = ttk.Button(input_buttons, text="Send", command=self.on_send_text)
        self.send_btn.pack(fill=tk.X, pady=(0, 6))

        # Mic button (single-shot listen)
        self.mic_btn = ttk.Button(input_buttons, text="🎤", width=3, command=self.single_listen_once)
        self.mic_btn.pack(fill=tk.X)

        # Panel for "Previous Chats" (listbox) on the rightmost side
        previous_panel = ttk.Frame(content_pane, width=300)
        previous_panel.pack_propagate(False)
        content_pane.add(previous_panel, weight=1)

        prev_label = ttk.Label(previous_panel, text="Previous Chats", font=("Helvetica", 12, "bold"))
        prev_label.pack(anchor=tk.W, padx=6, pady=(6, 0))

        # Listbox shows brief preview, click to populate console/input
        self.prev_listbox = tk.Listbox(previous_panel)
        self.prev_listbox.pack(fill=tk.BOTH, expand=True, padx=6, pady=6)
        self.prev_listbox.bind("<<ListboxSelect>>", self.on_select_prev_chat)

        # Footer tips
        footer_frame = ttk.Frame(main_container)
        footer_frame.pack(fill=tk.X, pady=(10, 0))
        tips = [
            "Try: 'Open Chrome'",
            "Try: 'What is the time?'",
            "Try: 'Search for AI news'",
            "Try: 'Excel mode on'"
        ]
        tip_var = tk.StringVar(value=tips[0])
        tip_label = ttk.Label(footer_frame, textvariable=tip_var, style='Status.TLabel')
        tip_label.pack()
        def cycle_tips():
            current = tip_var.get()
            next_index = (tips.index(current) + 1) % len(tips)
            tip_var.set(tips[next_index])
            self.root.after(5000, cycle_tips)
        self.root.after(5000, cycle_tips)

        # Animation and stats
        self.animate_title()
        self.animate_visualizer()
        self.update_system_stats()

    # ---------------- UI helper methods ----------------
    def animate_title(self):
        current_text = self.title_label.cget("text")
        if "PHOENIX AI" in current_text:
            new_text = current_text.replace("PHOENIX AI", "PHŒNIX AI")
        else:
            new_text = current_text.replace("PHŒNIX AI", "PHOENIX AI")
        self.title_label.config(text=new_text)
        self.root.after(1000, self.animate_title)

    def update_system_stats(self):
        try:
            self.cpu_var.set(f"CPU: {psutil.cpu_percent()}%")
            self.mem_var.set(f"Memory: {psutil.virtual_memory().percent}%")
        except Exception:
            pass
        self.root.after(2000, self.update_system_stats)

    def animate_visualizer(self):
        self.canvas.delete("all")
        width = self.canvas.winfo_width()
        height = self.canvas.winfo_height()
        if width > 1 and height > 1:
            for i in range(0, width, 6):
                amplitude = 15 + 20 * abs((time.time() * 6 + i) % 12 - 6)
                # keep color strings as before
                color = "#00ff00" if i % 18 < 6 else "#00ffff" if i % 18 < 12 else "#ff00ff"
                self.canvas.create_line(i, height / 2 - amplitude / 2, i, height / 2 + amplitude / 2,
                                        fill=color, width=2)
        self.root.after(80, self.animate_visualizer)

    def log_message(self, message, sender="Phoenix"):
        # Add to chat_history with timestamp
        ts = datetime.datetime.now().isoformat()
        self.chat_history.append((sender, message, ts))
        # Update console view
        self.console.config(state=tk.NORMAL)
        timestamp = datetime.datetime.now().strftime("%H:%M:%S")
        if sender == "Phoenix":
            tag = "ai"
            self.console.tag_config("ai", foreground="#00ff00")
        elif sender == "You":
            tag = "user"
            self.console.tag_config("user", foreground="#00ffff")
        else:
            tag = "system"
            self.console.tag_config("system", foreground="#ff5555")
        self.console.insert(tk.END, f"[{timestamp}] {sender}: {message}\n", tag)
        self.console.see(tk.END)
        self.console.config(state=tk.DISABLED)
        # Also update previous chats listbox (show first ~60 chars)
        preview = (message[:60] + "...") if len(message) > 60 else message
        display = f"{sender} - {preview} ({timestamp})"
        self.prev_listbox.insert(tk.END, display)

    def speak(self, text):
        # All AI replies go through this function to both log and speak
        self.log_message(text, "Phoenix")
        try:
            engine.say(text)
            engine.runAndWait()
        except Exception as e:
            # If TTS fails, log system message
            self.log_message(f"TTS Error: {str(e)}", "System")

    # ---------------- Authentication + listening ----------------
    def toggle_listening(self):
        if not self.authenticated:
            self.speak("Authentication required first.")
            return
        if not self.listening:
            self.listening = True
            self.listen_btn.config(text="Stop Listening")
            self.status_var.set("Listening...")
            threading.Thread(target=self.listen_loop, daemon=True).start()
        else:
            self.listening = False
            self.listen_btn.config(text="Start Listening")
            self.status_var.set("Authenticated")

    def authenticate(self):
        password_attempt = self.pass_entry.get()
        if password_attempt == PASSWORD:
            self.authenticated = True
            self.status_var.set("Authenticated")
            self.speak("Authentication successful. Phoenix is now activated. How can I assist you?")
            self.auth_btn.config(state='disabled', text="Authenticated")
            self.pass_entry.config(state='disabled')
        else:
            self.speak("Authentication failed. Please try again.")
            self.pass_entry.delete(0, tk.END)

    def listen_loop(self):
        while self.listening:
            try:
                with self.microphone as source:
                    self.status_var.set("Listening...")
                    audio = self.recognizer.listen(source, timeout=5, phrase_time_limit=8)
                self.status_var.set("Processing...")
                command = self.recognizer.recognize_google(audio).lower()
                self.log_message(command, "You")
                self.process_command(command)
            except sr.WaitTimeoutError:
                continue
            except sr.UnknownValueError:
                self.status_var.set("Could not understand")
            except Exception as e:
                self.log_message(f"Error: {str(e)}", "System")
                self.status_var.set("Error occurred")

    # Single-shot mic for the input area mic button
    def single_listen_once(self):
        if not self.authenticated:
            self.speak("Please authenticate first.")
            return
        threading.Thread(target=self._single_listen_thread, daemon=True).start()

    def _single_listen_thread(self):
        try:
            self.mic_btn.config(text="Listening...")
            with self.microphone as source:
                audio = self.recognizer.listen(source, timeout=5, phrase_time_limit=8)
            command = self.recognizer.recognize_google(audio).lower()
            # Put recognized text into the input box and send automatically
            self.input_text.delete("1.0", tk.END)
            self.input_text.insert(tk.END, command)
            self.log_message(command, "You")
            self.on_send_text()
        except sr.WaitTimeoutError:
            self.speak("No speech detected.")
        except sr.UnknownValueError:
            self.speak("Could not understand audio.")
        except Exception as e:
            self.log_message(f"Mic Error: {str(e)}", "System")
            self.speak("Microphone error.")
        finally:
            self.mic_btn.config(text="🎤")

    # ---------------- Input send/prev chat handling ----------------
    def on_send_text(self):
        text = self.input_text.get("1.0", tk.END).strip()
        if not text:
            return
        self.input_text.delete("1.0", tk.END)
        self.log_message(text, "You")
        # Route to same processing as voice commands
        self.process_command(text.lower())

    def on_select_prev_chat(self, event):
        # When user selects an item, show full text in a popup and option to re-send or copy
        selection = event.widget.curselection()
        if not selection:
            return
        idx = selection[0]
        # Map prev listbox index to chat_history index (they are in same order)
        try:
            sender, message, ts = self.chat_history[idx]
            # Show a small dialog with full message and actions
            popup = tk.Toplevel(self.root)
            popup.title("Chat Item")
            popup.geometry("500x300")
            lbl = ttk.Label(popup, text=f"{sender} at {ts}", font=("Segoe UI", 10, "bold"))
            lbl.pack(anchor=tk.W, padx=10, pady=(10, 0))
            txt = scrolledtext.ScrolledText(popup, height=10, wrap=tk.WORD)
            txt.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            txt.insert(tk.END, message)
            txt.config(state=tk.DISABLED)
            btn_frame = ttk.Frame(popup)
            btn_frame.pack(fill=tk.X, pady=(0, 10), padx=10)
            def resend():
                popup.destroy()
                # Put the message into input and process as user command
                self.input_text.delete("1.0", tk.END)
                self.input_text.insert(tk.END, message)
                self.on_send_text()
            ttk.Button(btn_frame, text="Resend", command=resend).pack(side=tk.LEFT)
            def copy_to_clipboard():
                self.root.clipboard_clear()
                self.root.clipboard_append(message)
                popup.destroy()
            ttk.Button(btn_frame, text="Copy", command=copy_to_clipboard).pack(side=tk.LEFT, padx=(6,0))
        except Exception as e:
            self.log_message(f"PrevChat access error: {str(e)}", "System")

    # ---------------- Command processing ----------------
    def process_command(self, command):
        if not command:
            return

        # PC power commands (require authentication)
        if "shutdown" in command or "shut down" in command:
            if not self.authenticated:
                self.speak("You must authenticate before shutting down the PC.")
                return
            # Ask for confirmation (non-blocking)
            if messagebox.askyesno("Confirm Shutdown", "Confirm shutdown the PC?"):
                self.speak("Shutting down the PC now.")
                try:
                    if os.name == 'nt':
                        subprocess.call(["shutdown", "/s", "/t", "5"])
                    elif sys.platform == 'darwin':
                        subprocess.call(["sudo", "shutdown", "-h", "now"])
                    else:
                        subprocess.call(["shutdown", "-h", "now"])
                except Exception as e:
                    self.speak(f"Failed to shutdown: {e}")
            return
        if "restart" in command or "reboot" in command:
            if not self.authenticated:
                self.speak("You must authenticate before restarting the PC.")
                return
            if messagebox.askyesno("Confirm Restart", "Confirm restart the PC?"):
                self.speak("Restarting the PC now.")
                try:
                    if os.name == 'nt':
                        subprocess.call(["shutdown", "/r", "/t", "5"])
                    elif sys.platform == 'darwin':
                        subprocess.call(["sudo", "shutdown", "-r", "now"])
                    else:
                        subprocess.call(["reboot"])
                except Exception as e:
                    self.speak(f"Failed to restart: {e}")
            return

        # Regular commands as before
        if "open " in command:
            app = command.replace("open", "").strip()
            self.open_application(app)
        elif command.startswith("search ") or command.startswith("search for "):
            query = command.replace("search", "").replace("for", "").strip()
            self.search_web(query)
        elif command.startswith("what is") or command.startswith("who is") or "tell me about" in command:
            query = command.replace("what is", "").replace("who is", "").replace("tell me about", "").strip()
            self.ask_gemini(query)
        elif "time" in command and len(command.split()) <= 3:
            self.speak(f"The current time is {datetime.datetime.now().strftime('%I:%M %p')}")
        elif "date" in command and len(command.split()) <= 3:
            self.speak(f"Today is {datetime.datetime.now().strftime('%A, %B %d, %Y')}")
        elif ("write to excel" in command) or ("excel" in command and "mode" not in command):
            self.handle_excel_command(command)
        elif "excel mode on" in command or "excel mode off" in command:
            # toggle regardless of phrasing
            self.toggle_excel_mode()
        elif "ask deepseek" in command or "deepseek" in command:
            query = command.replace("ask deepseek", "").replace("deepseek", "").strip()
            self.ask_deepseek(query)
        elif "wikipedia" in command:
            query = command.replace("wikipedia", "").replace("search", "").strip()
            self.search_wikipedia(query)
        elif "exit" in command or "quit" in command or "goodbye" in command:
            self.speak("Goodbye Master. Shutting down.")
            self.root.after(2000, self.root.destroy)
        else:
            # fallback to general assistant (Gemini)
            self.ask_gemini(command)

    # ---------------- App helpers ----------------
    def open_application(self, app_name):
        self.speak(f"Opening {app_name}")
        path = APP_PATHS.get(app_name, None)
        if path:
            if path.startswith("http"):
                webbrowser.open(path)
            else:
                try:
                    if os.name == 'nt':
                        os.startfile(path)
                    elif sys.platform == 'darwin':
                        subprocess.call(['open', path])
                    else:
                        subprocess.call(['xdg-open', path])
                except Exception as e:
                    self.speak(f"Sorry, I couldn't open {app_name}. Error: {str(e)}")
        else:
            webbrowser.open(f"https://{app_name}")

    def search_web(self, query):
        self.speak(f"Searching for {query}")
        webbrowser.open(f"https://www.google.com/search?q={query}")

    def search_wikipedia(self, query):
        try:
            result = wikipedia.summary(query, sentences=2)
            self.speak(f"According to Wikipedia: {result}")
        except wikipedia.exceptions.DisambiguationError:
            self.speak(f"Multiple results found for {query}. Please be more specific.")
        except wikipedia.exceptions.PageError:
            self.speak(f"Sorry, I couldn't find any information about {query}.")
        except Exception:
            self.speak("Sorry, I encountered an error while searching Wikipedia.")

    # ---------------- Gemini + Deepseek wrappers ----------------
    def ask_gemini(self, query):
        # Keep Gemini config exactly as provided. We only add logging+TTS (speak) and record response into history.
        try:
            headers = {'Content-Type': 'application/json'}
            data = {"contents": [{"parts": [{"text": query}]}]}
            response = requests.post(GEMINI_API_URL, headers=headers, json=data)
            response.raise_for_status()
            result = response.json()
            candidates = result.get('candidates', [])
            if candidates and 'content' in candidates[0]:
                parts = candidates[0]['content'].get('parts', [])
                if parts:
                    text = parts[0].get('text', "")
                    # Speak and also log
                    self.speak(text)
                else:
                    self.speak("No response from Gemini.")
            else:
                self.speak("No response from Gemini.")
        except Exception as e:
            self.speak("Error contacting Gemini API.")
            self.log_message(f"Gemini API Error: {str(e)}", "System")

    def ask_deepseek(self, query):
        try:
            headers = {'Content-Type': 'application/json', 'Authorization': f'Bearer {DEEPSEEK_API_KEY}'}
            data = {
                "model": "deepseek-chat",
                "messages": [{"role": "user", "content": query}],
                "temperature": 0.7,
                "max_tokens": 2000
            }
            response = requests.post(DEEPSEEK_API_URL, headers=headers, json=data)
            response.raise_for_status()
            result = response.json()
            text = result.get('choices', [{}])[0].get('message', {}).get('content', '')
            self.speak(text)
        except Exception as e:
            self.speak("Error contacting DeepSeek API.")
            self.log_message(f"DeepSeek API Error: {str(e)}", "System")

    # ---------------- Excel helpers ----------------
    def toggle_excel_mode(self):
        if not self.excel_mode:
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Create or Select Excel File"
            )
            if file_path:
                self.current_excel_file = file_path
                self.create_excel_file()
                self.excel_mode = True
                self.excel_btn.config(text="Excel Mode: On")
                self.speak("Excel mode activated. You can now give commands to write data.")
            else:
                self.speak("Excel mode not activated.")
        else:
            self.current_excel_file = None
            self.excel_mode = False
            self.excel_btn.config(text="Excel Mode: Off")
            self.speak("Excel mode deactivated.")

    def create_excel_file(self):
        if not os.path.exists(self.current_excel_file):
            wb = Workbook()
            ws = wb.active
            ws.title = "Data"
            ws['A1'] = "Date"
            ws['B1'] = "Time"
            ws['C1'] = "Command"
            ws['D1'] = "Data"
            wb.save(self.current_excel_file)

    def handle_excel_command(self, command):
        if not self.excel_mode:
            self.speak("Please activate Excel mode first.")
            return
        try:
            wb = openpyxl.load_workbook(self.current_excel_file)
            ws = wb.active
            next_row = ws.max_row + 1
            if "add" in command and "to excel" in command:
                data = command.replace("add", "").replace("to excel", "").strip()
                ws[f'A{next_row}'] = datetime.datetime.now().strftime("%Y-%m-%d")
                ws[f'B{next_row}'] = datetime.datetime.now().strftime("%H:%M:%S")
                ws[f'C{next_row}'] = "Add command"
                ws[f'D{next_row}'] = data
                wb.save(self.current_excel_file)
                self.speak(f"Added '{data}' to Excel.")
            else:
                self.speak("Specify what to add to Excel. Example: 'Add groceries to excel'")
        except Exception as e:
            self.speak("Error working with Excel.")
            self.log_message(f"Excel Error: {str(e)}", "System")

    def open_file_dialog(self):
        file_path = filedialog.askopenfilename()
        if file_path:
            try:
                if os.name == 'nt':
                    os.startfile(file_path)
                elif sys.platform == 'darwin':
                    subprocess.call(['open', file_path])
                else:
                    subprocess.call(['xdg-open', file_path])
                self.speak(f"Opening {os.path.basename(file_path)}")
            except Exception as e:
                self.speak(f"Couldn't open file. Error: {str(e)}")

    def quick_action(self, action):
        if action == "time":
            self.speak(f"The current time is {datetime.datetime.now().strftime('%I:%M %p')}")
        elif action == "date":
            self.speak(f"Today is {datetime.datetime.now().strftime('%A, %B %d, %Y')}")
        elif action == "weather":
            self.speak("Weather information is not configured yet. Please set up a weather API.")
        elif action == "news":
            webbrowser.open("https://news.google.com")
            self.speak("Opening Google News")
        elif action == "email":
            webbrowser.open("https://gmail.com")
            self.speak("Opening Gmail")
        elif action == "calendar":
            webbrowser.open("https://calendar.google.com")
            self.speak("Opening Google Calendar")


def main():
    root = tk.Tk()
    app = PhoenixAI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
